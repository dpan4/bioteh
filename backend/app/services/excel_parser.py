"""Excel-парсер и экспортер задач диаграммы Гантта (openpyxl).

Модуль отвечает ТОЛЬКО за преобразование данных:
- parse_excel(): байты .xlsx -> list[RawTask] (БЕЗ дат — даты вычисляет
  исключительно backend/app/services/scheduler.py, system.md, инвариант 1);
- generate_excel(): list[GanttTask] -> байты .xlsx с готовым расписанием (8 колонок);
- generate_template_excel(): -> байты .xlsx с эталонным шаблоном для заполнения (5 колонок).

Формат входного файла — строго 5 колонок (регистр и пробелы игнорируются):
  Задача / Title / Название
  Описание / Description
  Исполнитель / Assignee
  Длительность / Duration
  Предшественники / Predecessors

ID генерируется автоматически: task-1, task-2, ... по порядку строк.
Предшественники: числовые номера (1, 2 -> task-1, task-2) или строковые ID.

Формат выходного файла экспорта — 8 колонок:
  ID | Задача | Описание | Исполнитель | Длительность | Дата начала | Дата окончания | Предшественники
Дата окончания — живая формула Excel (= Дата начала + Длительность).

Формат шаблона для заполнения — строго 5 колонок:
  Задача | Описание | Исполнитель | Длительность (дни) | Предшественники
БЕЗ колонок ID и дат (даты рассчитает бэкенд при импорте).
"""

import io
import re
from datetime import date as date_type
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import ValidationError

from app.schemas.task import GanttTask, RawTask

# ── Маппинг колонок (строго 5, без ID) ──────────────────────────────────────

_COLUMN_ALIASES: dict[str, str] = {
    "задача": "title",
    "title": "title",
    "название": "title",
    "описание": "description",
    "description": "description",
    "исполнитель": "assignee",
    "assignee": "assignee",
    "длительность": "duration_days",
    "длительность (дни)": "duration_days",
    "duration": "duration_days",
    "duration_days": "duration_days",
    "предшественники": "predecessors",
    "predecessors": "predecessors",
}

_REQUIRED_FIELDS: dict[str, str] = {
    "title": '"Задача" / "Title" / "Название"',
    "duration_days": '"Длительность" / "Duration"',
}

# ── Колонки выходного Excel (8 колонок) ─────────────────────────────────────
# A=ID, B=Задача, C=Описание, D=Исполнитель, E=Длительность,
# F=Дата начала, G=Дата окончания, H=Предшественники

_EXPORT_COLUMNS: list[tuple[str, int]] = [
    ("ID", 10),
    ("Задача", 30),
    ("Описание", 45),
    ("Исполнитель", 20),
    ("Длительность", 14),
    ("Дата начала", 14),
    ("Дата окончания", 14),
    ("Предшественники", 25),
]

# Индексы колонок для формул (1-based)
_COL_START = 6   # F — Дата начала
_COL_DURATION = 5  # E — Длительность
_COL_END = 7       # G — Дата окончания

# ── Вспомогательные функции ──────────────────────────────────────────────────


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip().lower()


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_duration(value: Any, row_number: int) -> int:
    text = _cell_to_str(value)
    if text == "":
        raise ValueError(
            f"Строка {row_number}: не заполнена длительность задачи. "
            "Укажите целое число рабочих дней (не меньше 1)."
        )
    try:
        numeric = float(text.replace(",", "."))
    except (TypeError, ValueError):
        raise ValueError(
            f"Строка {row_number}: длительность '{text}' не является числом. "
            "Укажите целое число рабочих дней (не меньше 1)."
        )
    if not numeric.is_integer():
        raise ValueError(
            f"Строка {row_number}: длительность '{text}' должна быть целым "
            "числом рабочих дней."
        )
    duration = int(numeric)
    if duration < 1:
        raise ValueError(
            f"Строка {row_number}: длительность должна быть не меньше 1 дня, "
            f"получено {duration}."
        )
    return duration


def _parse_predecessors(value: Any, row_number: int) -> list[str]:
    """Разбирает предшественников: числовые номера строк или строковые ID.

    Поддерживаются разделители запятая и точка с запятой.

    Числовые значения (1, 2, 3) преобразуются в task-N:
      "1"       -> ["task-1"]
      "1, 2"    -> ["task-1", "task-2"]
      "1; 2"    -> ["task-1", "task-2"]

    Строковые ID оставляются как есть:
      "task-1, task-2" -> ["task-1", "task-2"]

    Args:
        value: Сырое значение ячейки предшественников.
        row_number: Номер строки (для сообщений об ошибках).

    Returns:
        Список ID задач-предшественников (возможно пустой).
    """
    text = _cell_to_str(value)
    if text == "":
        return []

    parts = [p.strip() for p in re.split(r"[;,]", text) if p.strip()]
    result: list[str] = []
    for part in parts:
        if re.match(r"^\d+$", part):
            num = int(part)
            if num < 1:
                raise ValueError(
                    f"Строка {row_number}: номер предшественника '{part}' "
                    "должен быть положительным числом."
                )
            if num >= row_number:
                raise ValueError(
                    f"Строка {row_number}: номер предшественника '{part}' "
                    "ссылается на строку ({num}) не ранее текущей ({row_number}). "
                    "Предшественник должен быть выше по списку."
                )
            result.append(f"task-{num}")
        else:
            result.append(part)
    return result


# ── Парсер импорта ───────────────────────────────────────────────────────────


def parse_excel(file_bytes: bytes) -> list[RawTask]:
    """Читает .xlsx-файл со строго 5 колонками и возвращает список RawTask.

    ID генерируется автоматически: первая строка данных -> task-1,
    вторая -> task-2, и т.д.

    Ожидаемые колонки (регистр и лишние пробелы игнорируются):
      Задача / Title / Название           (обязательная)
      Описание / Description              (опциональная, default "")
      Исполнитель / Assignee              (опциональная, default "")
      Длительность / Duration             (обязательная, int >= 1)
      Предшественники / Predecessors      (опциональная, default [])

    Предшественники: числа (1, 2) -> task-1, task-2; строки -> как есть.

    Даты НЕ вычисляются — это делает только scheduler.py (инвариант 1).

    Args:
        file_bytes: Содержимое Excel-файла (.xlsx) в виде байтов.

    Returns:
        Список валидированных RawTask в порядке строк файла.

    Raises:
        ValueError: Файл повреждён, отсутствуют обязательные колонки,
            невалидные значения, пустой файл.
    """
    try:
        workbook = load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
    except Exception as exc:
        raise ValueError(
            "Не удалось открыть Excel-файл: файл повреждён или имеет "
            f"неподдерживаемый формат (ожидается .xlsx). Детали: {exc}"
        ) from exc

    try:
        worksheet = workbook.active
        if worksheet is None:
            raise ValueError("Excel-файл не содержит ни одного листа.")

        rows = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            raise ValueError(
                "Excel-файл пуст: отсутствует строка с названиями колонок."
            )

        column_map: dict[int, str] = {}
        for index, raw_header in enumerate(header_row):
            field = _COLUMN_ALIASES.get(_normalize_header(raw_header))
            if field is not None and field not in column_map.values():
                column_map[index] = field

        mapped_fields = set(column_map.values())
        missing = [
            label
            for field, label in _REQUIRED_FIELDS.items()
            if field not in mapped_fields
        ]
        if missing:
            raise ValueError(
                "В Excel-файле отсутствуют обязательные колонки: "
                f"{', '.join(missing)}. Проверьте первую строку файла."
            )

        tasks: list[RawTask] = []
        task_index = 0

        for row_number, row in enumerate(rows, start=2):
            values: dict[str, Any] = {
                field: row[index] if index < len(row) else None
                for index, field in column_map.items()
            }

            title = _cell_to_str(values.get("title"))
            other_cells_empty = all(
                _cell_to_str(values.get(field)) == ""
                for field in ("description", "assignee", "duration_days",
                              "predecessors")
            )
            if title == "" and other_cells_empty:
                continue
            if title == "":
                raise ValueError(
                    f"Строка {row_number}: не заполнено название задачи "
                    '(колонка "Задача" / "Title" / "Название").'
                )

            task_index += 1
            task_id = f"task-{task_index}"

            try:
                task = RawTask(
                    id=task_id,
                    title=title,
                    description=_cell_to_str(values.get("description")),
                    assignee=_cell_to_str(values.get("assignee")),
                    duration_days=_parse_duration(
                        values.get("duration_days"), row_number
                    ),
                    predecessors=_parse_predecessors(
                        values.get("predecessors"), task_index + 1
                    ),
                )
            except ValidationError as exc:
                raise ValueError(
                    f"Строка {row_number}: данные задачи не прошли валидацию. "
                    f"Детали: {exc}"
                ) from exc

            tasks.append(task)

        if not tasks:
            raise ValueError(
                "Excel-файл не содержит ни одной задачи: заполните строки "
                "под шапкой таблицы."
            )

        return tasks
    finally:
        workbook.close()


# ── Экспортер ────────────────────────────────────────────────────────────────


def generate_excel(tasks: list[GanttTask]) -> bytes:
    """Формирует .xlsx-файл с расписанием (8 колонок, формулы, оформление).

    Колонки:
      A: ID
      B: Задача
      C: Описание
      D: Исполнитель
      E: Длительность
      F: Дата начала   (ISO YYYY-MM-DD)
      G: Дата окончания (живая формула = F{row} + E{row})
      H: Предшественники

    Args:
        tasks: Список задач с датами, рассчитанными DAG Scheduler.

    Returns:
        Содержимое Excel-файла в виде байтов.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "График проекта"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )

    for column_index, (header, width) in enumerate(_EXPORT_COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.freeze_panes = "A2"

    for row_index, task in enumerate(tasks, start=2):
        worksheet.cell(row=row_index, column=1, value=task.id)
        worksheet.cell(row=row_index, column=2, value=task.title)
        worksheet.cell(row=row_index, column=3, value=task.description)
        worksheet.cell(row=row_index, column=4, value=task.assignee)
        worksheet.cell(row=row_index, column=5, value=task.duration_days)

        start_cell = worksheet.cell(
            row=row_index, column=_COL_START, value=task.start_date
        )
        start_cell.number_format = "YYYY-MM-DD"

        end_formula = (
            f"={get_column_letter(_COL_START)}{row_index}"
            f"+{get_column_letter(_COL_DURATION)}{row_index}"
        )
        end_cell = worksheet.cell(
            row=row_index, column=_COL_END, value=end_formula
        )
        end_cell.number_format = "YYYY-MM-DD"

        pred_text = ", ".join(task.predecessors)
        worksheet.cell(row=row_index, column=8, value=pred_text)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ── Генератор эталонного шаблона ─────────────────────────────────────────────


def generate_template_excel() -> bytes:
    """Формирует эталонный шаблон .xlsx для заполнения проекта (5 колонок).

    Шаблон содержит:
    - Строго 5 колонок: Задача, Описание, Исполнитель, Длительность (дни), Предшественники.
    - БЕЗ колонок ID, Дата начала, Дата окончания (даты рассчитает бэкенд при импорте).
    - 5 наглядных демо-строк с реалистичными данными.
    - Предшественники — простые номера строк (1, 2, 3, 4), НЕ task-N.
    - Стилизованную шапку (темно-синяя заливка, белый жирный текст).
    - Автоматическую подгонку ширины колонок.

    Этот файл предназначен для скачивания пользователями через кнопку
    «Скачать шаблон» или MCP-инструмент get_excel_template.

    Returns:
        Содержимое Excel-файла в виде байтов.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Шаблон проекта"

    # ── Определение колонок шаблона (5 колонок, БЕЗ ID и дат) ────────────────
    template_columns: list[tuple[str, int]] = [
        ("Задача", 35),
        ("Описание", 50),
        ("Исполнитель", 20),
        ("Длительность (дни)", 20),
        ("Предшественники", 20),
    ]

    # ── Стилизация шапки ──────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )

    for column_index, (header, width) in enumerate(template_columns, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.freeze_panes = "A2"

    # ── Демо-данные (5 строк, демонстрирующие параллельность и слияние) ──────
    # Граф зависимостей (наглядный пример для пользователей):
    #
    #   A[1: Анализ требований, 5 дней] ──────────┐
    #   (старт День 1, без предшественников)      │
    #                                              ▼
    #   B[2: Исследование конкурентов, 3 дня]  C[3: Проектирование БД, 4 дня]
    #   (старт День 1, ПАРАЛЛЕЛЬНО с A!)       (после A → старт День 6)
    #          │                                   │
    #          ▼                                   │
    #   D[4: UI-макеты, 4 дня]  ◄─────────────────┘
    #   (после B → старт День 4, ПАРАЛЛЕЛЬНО с C!)
    #          │                                   │
    #          └────────────────┬──────────────────┘
    #                           ▼
    #            E[5: Интеграционное тестирование, 3 дня]
    #            (после max(C, D) → старт День 10)
    #
    demo_tasks = [
        {
            "title": "Анализ требований",
            "description": "Сбор бизнес-требований (корневая задача, старт в День 1)",
            "assignee": "Анна",
            "duration": 5,
            "predecessors": "",
        },
        {
            "title": "Исследование конкурентов",
            "description": "Анализ аналогов (идет параллельно со строкой 1!)",
            "assignee": "Игорь",
            "duration": 3,
            "predecessors": "",
        },
        {
            "title": "Проектирование БД",
            "description": "Схема PostgreSQL (строго после Анализа)",
            "assignee": "Дмитрий",
            "duration": 4,
            "predecessors": "1",
        },
        {
            "title": "Разработка UI-макетов",
            "description": "Дизайн интерфейса (после Исследования, параллельно с БД!)",
            "assignee": "Елена",
            "duration": 4,
            "predecessors": "2",
        },
        {
            "title": "Интеграционное тестирование",
            "description": "Проверка API и UI (ждёт завершения И БД, И макетов)",
            "assignee": "Анна",
            "duration": 3,
            "predecessors": "3, 4",
        },
    ]

    for row_index, task in enumerate(demo_tasks, start=2):
        worksheet.cell(row=row_index, column=1, value=task["title"])
        worksheet.cell(row=row_index, column=2, value=task["description"])
        worksheet.cell(row=row_index, column=3, value=task["assignee"])
        worksheet.cell(row=row_index, column=4, value=task["duration"])
        worksheet.cell(row=row_index, column=5, value=task["predecessors"])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()